#!/usr/bin/env python3
"""
Extraction one-shot des grilles de coûts ROD → ``data/couts.xlsx``.

Source (hors runtime, lecture seule)
------------------------------------
``../archive/sources/raw/ROD - Simulateurs + détail des coûts.xlsx``

Le classeur source contient des **formules** ; on l'ouvre avec
``openpyxl`` en ``data_only=True`` pour récupérer les **valeurs calculées**
(nécessite que le fichier ait été ouvert/sauvé une fois dans Excel/LibreOffice
si les caches de formules sont vides).

Feuilles produites
------------------
* ``resume`` — synthèse par solution (simply / liberty / connected)
* ``couts_technos`` — matériel, licences, frais ad hoc (format long)
* ``couts_annexes`` — électricité + personnel
* ``couts_agencement`` — m linéaires × classic/premium/bespoke
* ``revenus_mix_marges`` / ``revenus_impact_to`` — références revenus
* ``meta`` — provenance

Usage
-----
    python extract_couts.py
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
ARCHIVE_RAW = ROOT.parent / "archive" / "sources" / "raw"
OUT = ROOT / "data" / "couts.xlsx"


def _find_source() -> Path:
    for p in ARCHIVE_RAW.iterdir():
        if p.suffix.lower() == ".xlsx" and "Simulateur" in p.name and "cout" in p.name.lower().replace("û", "u").replace("ô", "o"):
            return p
        if p.suffix.lower() == ".xlsx" and "Simulateur" in p.name:
            return p
    # fallback glob
    cands = list(ARCHIVE_RAW.glob("*Simulateurs*.xlsx"))
    if not cands:
        raise FileNotFoundError(f"Fichier coûts introuvable dans {ARCHIVE_RAW}")
    return cands[0]


def _num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        t = v.strip().replace(",", ".")
        if t.lower() in {"", "-", "–", "inclus", "n/a"}:
            return None
        try:
            return float(t)
        except ValueError:
            return None
    return None


def extract_technos(ws) -> pd.DataFrame:
    """COUTS - TECHNOS → long format par solution / rubrique / qty."""
    # blocks: (solution, equipment_label_col, qty_col, unit_col, total_col, buy_col, lease_or_mensuel_col, mode)
    # SIMPLY: scanner C6 name, qty C, unit D, BUY E, MENSUEL F
    rows: list[dict[str, Any]] = []

    blocks = [
        # simply scanner
        dict(solution="simply", equip="SCANNER", r0=8, r1=11, qty=3, unit=4, total=5, mensuel=6, mode="buy_only"),
        dict(solution="simply", equip="VITRINE", r0=15, r1=18, qty=3, unit=4, total=5, mensuel=6, mode="buy_only"),
        dict(solution="simply", equip="LICENCE_LOGICIELLE", r0=22, r1=22, qty=3, unit=4, total=5, mensuel=6, mode="mensuel"),
        dict(solution="simply", equip="FRAIS_AD_HOC", r0=26, r1=26, qty=3, unit=4, total=5, mensuel=6, mode="one_shot"),
        # liberty
        dict(solution="liberty", equip="CAISSE", r0=8, r1=11, qty=10, unit=11, total=12, buy=13, lease=14, mode="buy_or_lease"),
        dict(solution="liberty", equip="VITRINE", r0=15, r1=18, qty=10, unit=11, total=12, buy=13, lease=14, mode="buy_or_lease"),
        dict(solution="liberty", equip="LICENCE_LOGICIELLE", r0=22, r1=22, qty=10, unit=11, total=13, mensuel=14, mode="mensuel"),
        dict(solution="liberty", equip="FRAIS_AD_HOC", r0=26, r1=26, qty=10, unit=11, total=13, mensuel=14, mode="one_shot"),
        # connected
        dict(solution="connected", equip="FRIGO_FROID", r0=8, r1=11, qty=18, unit=19, total=20, buy=21, lease=22, mode="buy_or_lease"),
        dict(solution="connected", equip="FRIGO_AMBIANT", r0=15, r1=18, qty=18, unit=19, total=20, buy=21, lease=22, mode="buy_or_lease"),
        dict(solution="connected", equip="LICENCE_LOGICIELLE", r0=22, r1=22, qty=18, unit=19, total=21, mensuel=22, mode="mensuel_inclus"),
        dict(solution="connected", equip="FRAIS_AD_HOC", r0=26, r1=26, qty=18, unit=19, total=21, mensuel=22, mode="one_shot"),
    ]

    for b in blocks:
        for r in range(b["r0"], b["r1"] + 1):
            qty = _num(ws.cell(r, b["qty"]).value)
            unit = ws.cell(r, b["unit"]).value
            if qty is None and unit is None:
                continue
            row: dict[str, Any] = {
                "rubrique": "techno",
                "solution": b["solution"],
                "equipement": b["equip"],
                "quantite": int(qty) if qty is not None else None,
                "unite": str(unit).strip() if unit else None,
                "mode_acquisition": b["mode"],
                "duree_amort_mois": 60,
                "cout_total": None,
                "cout_buy": None,
                "cout_lease_mensuel": None,
                "cout_mensuel": None,
            }
            if "total" in b:
                row["cout_total"] = _num(ws.cell(r, b["total"]).value)
            if "buy" in b:
                row["cout_buy"] = _num(ws.cell(r, b["buy"]).value)
            if "lease" in b:
                row["cout_lease_mensuel"] = _num(ws.cell(r, b["lease"]).value)
            if "mensuel" in b:
                mv = ws.cell(r, b["mensuel"]).value
                if isinstance(mv, str) and "inclus" in mv.lower():
                    row["cout_mensuel"] = None
                    row["note"] = "inclus"
                else:
                    row["cout_mensuel"] = _num(mv)
            # simply: total col is BUY total
            if b["mode"] == "buy_only":
                row["cout_buy"] = row["cout_total"]
            rows.append(row)

    # totals row per solution
    for sol, tcol, mcol in (
        ("simply", 5, 6),
        ("liberty", 13, 14),
        ("connected", 21, 22),
    ):
        rows.append(
            {
                "rubrique": "techno",
                "solution": sol,
                "equipement": "TOTAL",
                "quantite": None,
                "unite": None,
                "mode_acquisition": "total",
                "duree_amort_mois": 60,
                "cout_total": _num(ws.cell(28, tcol).value),
                "cout_buy": _num(ws.cell(28, tcol).value) if sol != "liberty" else _num(ws.cell(28, 13).value),
                "cout_lease_mensuel": _num(ws.cell(28, mcol).value) if sol != "simply" else None,
                "cout_mensuel": _num(ws.cell(28, mcol).value),
            }
        )
    return pd.DataFrame(rows)


def extract_annexes(ws) -> pd.DataFrame:
    """COUTS - ANNEXES : electricité (scan/caisse/frigo + vitrine) + personnel."""
    rows: list[dict[str, Any]] = []
    # solutions column starts
    sols = [
        ("simply", 3, 4, 5, 6),
        ("liberty", 9, 10, 11, 12),
        ("connected", 15, 16, 17, 18),
    ]
    # block 1 electricity primary equipment rows 8-11
    for sol, q, u, tot, m in sols:
        for r in range(8, 12):
            qty = _num(ws.cell(r, q).value)
            unit = ws.cell(r, u).value
            if qty is None:
                continue
            rows.append(
                {
                    "rubrique": "annexe",
                    "sous_rubrique": "electricite",
                    "solution": sol,
                    "equipement": str(unit).strip() if unit else None,
                    "quantite": int(qty),
                    "duree_amort_mois": 60,
                    "cout_total": _num(ws.cell(r, tot).value),
                    "cout_mensuel": _num(ws.cell(r, m).value),
                }
            )
        # vitrine / second elec 15-18
        for r in range(15, 19):
            qty = _num(ws.cell(r, q).value)
            unit = ws.cell(r, u).value
            if qty is None:
                continue
            rows.append(
                {
                    "rubrique": "annexe",
                    "sous_rubrique": "electricite",
                    "solution": sol,
                    "equipement": str(unit).strip() if unit else "vitr.",
                    "quantite": int(qty),
                    "duree_amort_mois": 60,
                    "cout_total": _num(ws.cell(r, tot).value),
                    "cout_mensuel": _num(ws.cell(r, m).value),
                }
            )
        # personnel row 22
        qty = _num(ws.cell(22, q).value)
        unit = ws.cell(22, u).value
        rows.append(
            {
                "rubrique": "personnel",
                "sous_rubrique": "personnel",
                "solution": sol,
                "equipement": str(unit).strip() if unit else "staff",
                "quantite": int(qty) if qty else 1,
                "duree_amort_mois": 60,
                "cout_total": _num(ws.cell(22, tot).value),
                "cout_mensuel": _num(ws.cell(22, m).value),
            }
        )
        # total row 24
        rows.append(
            {
                "rubrique": "annexe",
                "sous_rubrique": "total_annexes_incl_personnel",
                "solution": sol,
                "equipement": "TOTAL",
                "quantite": None,
                "duree_amort_mois": 60,
                "cout_total": _num(ws.cell(24, tot).value),
                "cout_mensuel": _num(ws.cell(24, m).value),
            }
        )
    return pd.DataFrame(rows)


def extract_agencement(ws) -> pd.DataFrame:
    """COUTS - AGENCEMENT : m linéaire × disposition classic/premium/bespoke."""
    rows: list[dict[str, Any]] = []
    # (solution, qty_col, unit_col, classic_tot, classic_m, premium_tot, premium_m, bespoke_tot, bespoke_m)
    blocks = [
        ("simply", 3, 4, 5, 6, 7, 8, 9, 10),
        ("liberty", 13, 14, 15, 16, 17, 18, 19, 20),
        ("connected", 23, 24, 25, 26, 27, 28, 29, 30),
    ]
    for sol, q, u, ct, cm, pt, pm, bt, bm in blocks:
        for r in range(8, 38):
            qty = _num(ws.cell(r, q).value)
            unit = ws.cell(r, u).value
            if qty is None:
                continue
            for disp, tot_c, mens_c in (
                ("classic", ct, cm),
                ("premium", pt, pm),
                ("bespoke", bt, bm),
            ):
                rows.append(
                    {
                        "rubrique": "agencement",
                        "solution": sol,
                        "disposition": disp,
                        "metres_lineaires": int(qty),
                        "unite": str(unit).strip() if unit else "m",
                        "duree_amort_mois": 84,
                        "cout_total": _num(ws.cell(r, tot_c).value),
                        "cout_mensuel": _num(ws.cell(r, mens_c).value),
                    }
                )
    return pd.DataFrame(rows)


def extract_mix_marges(ws) -> pd.DataFrame:
    rows = []
    # simply cols 3-5, liberty 8-10, connected 13-15
    for sol, c_fb, c_nfb, c_tot in (
        ("simply", 3, 4, 5),
        ("liberty", 8, 9, 10),
        ("connected", 13, 14, 15),
    ):
        mix_fb = _num(ws.cell(7, c_fb).value)
        mix_nfb = _num(ws.cell(7, c_nfb).value)
        marge_fb = _num(ws.cell(8, c_fb).value)
        marge_nfb = _num(ws.cell(8, c_nfb).value)
        marge_pond = _num(ws.cell(8, c_tot).value)
        rows.append(
            {
                "solution": sol,
                "mix_f_b": mix_fb,
                "mix_n_f_b": mix_nfb,
                "marge_f_b": marge_fb,
                "marge_n_f_b": marge_nfb,
                "marge_ponderee": marge_pond,
            }
        )
    # moyenne row
    rows.append(
        {
            "solution": "simply_moyenne",
            "mix_f_b": None,
            "mix_n_f_b": None,
            "marge_f_b": None,
            "marge_n_f_b": None,
            "marge_ponderee": _num(ws.cell(18, 5).value),
        }
    )
    rows.append(
        {
            "solution": "liberty_moyenne",
            "marge_ponderee": _num(ws.cell(18, 10).value),
        }
    )
    rows.append(
        {
            "solution": "connected_moyenne",
            "marge_ponderee": _num(ws.cell(18, 15).value),
        }
    )
    return pd.DataFrame(rows)


def extract_impact_to(ws) -> pd.DataFrame:
    """Impact TO et CA pilotes."""
    rows = []
    # simply: hotel IBB NICE row 8
    rows.append(
        {
            "solution": "simply",
            "hotel_ref": ws.cell(8, 2).value,
            "to_moyen": _num(ws.cell(8, 3).value),
            "ca_ht_f_b": _num(ws.cell(8, 4).value),
            "ca_ht_n_f_b": _num(ws.cell(8, 5).value),
            "ca_ht_total": _num(ws.cell(8, 6).value),
            "ca_ttc_f_b": _num(ws.cell(8, 7).value),
            "ca_ttc_n_f_b": _num(ws.cell(8, 8).value),
            "ca_ttc_total": _num(ws.cell(8, 9).value),
            "type": "pilote",
        }
    )
    rows.append(
        {
            "solution": "simply",
            "hotel_ref": "IMPACT_1pct_TO",
            "to_moyen": _num(ws.cell(12, 3).value),
            "ca_ht_f_b": _num(ws.cell(12, 4).value),
            "ca_ht_n_f_b": _num(ws.cell(12, 5).value),
            "ca_ht_total": _num(ws.cell(12, 6).value),
            "ca_ttc_f_b": _num(ws.cell(12, 7).value),
            "ca_ttc_n_f_b": _num(ws.cell(12, 8).value),
            "ca_ttc_total": _num(ws.cell(12, 9).value),
            "type": "impact_1pct",
        }
    )
    # liberty hotels
    for r, typ in ((8, "pilote"), (9, "pilote"), (10, "moyenne"), (12, "impact_1pct")):
        href = ws.cell(r, 11).value
        if href is None and typ != "impact_1pct":
            continue
        rows.append(
            {
                "solution": "liberty",
                "hotel_ref": href if href else "IMPACT_1pct_TO",
                "to_moyen": _num(ws.cell(r, 12).value),
                "ca_ht_f_b": _num(ws.cell(r, 13).value),
                "ca_ht_n_f_b": _num(ws.cell(r, 14).value),
                "ca_ht_total": _num(ws.cell(r, 15).value),
                "type": typ,
            }
        )
    return pd.DataFrame(rows)


def extract_summary(technos: pd.DataFrame, annexes: pd.DataFrame, agencement: pd.DataFrame) -> pd.DataFrame:
    """Vue synthétique par solution / rubrique (totaux de référence)."""
    rows = []
    for sol in ("simply", "liberty", "connected"):
        t = technos[(technos["solution"] == sol) & (technos["equipement"] == "TOTAL")]
        a = annexes[
            (annexes["solution"] == sol)
            & (annexes["sous_rubrique"] == "total_annexes_incl_personnel")
        ]
        p = annexes[(annexes["solution"] == sol) & (annexes["rubrique"] == "personnel")]
        # agencement ref at 6m classic (typical)
        ag = agencement[
            (agencement["solution"] == sol)
            & (agencement["disposition"] == "classic")
            & (agencement["metres_lineaires"] == 6)
        ]
        rows.append(
            {
                "solution": sol,
                "cout_techno_total": float(t["cout_total"].iloc[0]) if len(t) and pd.notna(t["cout_total"].iloc[0]) else None,
                "cout_techno_mensuel": float(t["cout_mensuel"].iloc[0]) if len(t) and pd.notna(t["cout_mensuel"].iloc[0]) else None,
                "cout_annexes_total": float(a["cout_total"].iloc[0]) if len(a) and pd.notna(a["cout_total"].iloc[0]) else None,
                "cout_annexes_mensuel": float(a["cout_mensuel"].iloc[0]) if len(a) and pd.notna(a["cout_mensuel"].iloc[0]) else None,
                "cout_personnel_total": float(p["cout_total"].iloc[0]) if len(p) and pd.notna(p["cout_total"].iloc[0]) else None,
                "cout_personnel_mensuel": float(p["cout_mensuel"].iloc[0]) if len(p) and pd.notna(p["cout_mensuel"].iloc[0]) else None,
                "cout_agencement_classic_6m_total": float(ag["cout_total"].iloc[0]) if len(ag) else None,
                "cout_agencement_classic_6m_mensuel": float(ag["cout_mensuel"].iloc[0]) if len(ag) else None,
            }
        )
    return pd.DataFrame(rows)


def main() -> int:
    src = _find_source()
    print(f"Source: {src}")
    wb = load_workbook(src, data_only=True)

    technos = extract_technos(wb["COUTS - TECHNOS"])
    annexes = extract_annexes(wb["COUTS - ANNEXES"])
    agencement = extract_agencement(wb["COUTS - AGENCEMENT"])
    mix = extract_mix_marges(wb["REVENUS - MIX & MARGES"])
    impact = extract_impact_to(wb["REVENUS - IMPACT TO"])
    summary = extract_summary(technos, annexes, agencement)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="resume")
        technos.to_excel(writer, index=False, sheet_name="couts_technos")
        annexes.to_excel(writer, index=False, sheet_name="couts_annexes")
        agencement.to_excel(writer, index=False, sheet_name="couts_agencement")
        mix.to_excel(writer, index=False, sheet_name="revenus_mix_marges")
        impact.to_excel(writer, index=False, sheet_name="revenus_impact_to")
        # meta
        meta = pd.DataFrame(
            [
                {"key": "source", "value": str(src.name)},
                {"key": "note", "value": "Valeurs calculées (data_only) depuis le fichier ROD Simulateurs"},
                {"key": "solutions", "value": "simply, liberty, connected"},
                {"key": "rubriques", "value": "techno, annexe (electricite), personnel, agencement (classic/premium/bespoke)"},
            ]
        )
        meta.to_excel(writer, index=False, sheet_name="meta")

    print(f"→ {OUT}")
    print(f"  technos={len(technos)} annexes={len(annexes)} agencement={len(agencement)}")
    print(summary.to_string(index=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
