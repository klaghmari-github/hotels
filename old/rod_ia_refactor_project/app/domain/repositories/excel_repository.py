from dataclasses import dataclass
from pathlib import Path
from typing import Any
import openpyxl

@dataclass
class ExcelCellFormula:
    workbook: str
    sheet: str
    cell: str
    formula: str
    value: Any = None
    comment: str | None = None

class ExcelRuleRepository:
    """Inventaire technique des feuilles, formules et commentaires Excel."""
    def __init__(self, workbook_path: str | Path):
        self.workbook_path = Path(workbook_path)

    def list_sheets(self) -> list[str]:
        wb = openpyxl.load_workbook(self.workbook_path, data_only=False, read_only=True)
        return wb.sheetnames

    def extract_formulas(self) -> list[ExcelCellFormula]:
        wb_formula = openpyxl.load_workbook(self.workbook_path, data_only=False, read_only=False)
        wb_values = openpyxl.load_workbook(self.workbook_path, data_only=True, read_only=False)
        out: list[ExcelCellFormula] = []
        for ws in wb_formula.worksheets:
            ws_values = wb_values[ws.title]
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith('='):
                        value = ws_values[c.coordinate].value
                        comment = c.comment.text if c.comment else None
                        out.append(ExcelCellFormula(self.workbook_path.name, ws.title, c.coordinate, c.value, value, comment))
        return out

    def extract_comments(self) -> list[dict]:
        wb = openpyxl.load_workbook(self.workbook_path, data_only=False, read_only=False)
        out = []
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if c.comment:
                        out.append({
                            "workbook": self.workbook_path.name,
                            "sheet": ws.title,
                            "cell": c.coordinate,
                            "comment": c.comment.text,
                            "value": c.value,
                        })
        return out
