"""Extraction du Récapitulatif ROD (wide → long → features ``d_recap_*``)."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from rod_ia.domain.repositories.identity_registry import HotelIdentityRegistry
from rod_ia.domain.services.ml_column_naming import MLColumnNaming

# Colonnes hôtel du fichier « Récapitulatif … ROD (2).xlsx » → hotel_id canonique
RECAP_COLUMN_TO_HOTEL_ID: dict[str, str] = {
    "NICE": "ibis-budget-nice",
    "STRASBOURG": "ibis-budget-strasbourg",
    "PARIS CDG": "ibis-styles-roissy-cdg",
    "MEGEVE": "novotel-megeve",
    "TOUR EIFFEL": "novotel-paris-tour-eiffel",
    "MONTMARTRE": "mercure-montmartre",
    "BOULOGNE": "mercure-boulogne",
}

MISSING_TOKENS = frozenset({"", "?", "-", "nan", "none", "n/a"})

# Identifiants récap (section « PAGE DE CONNEXION / ID ») — pas des features ML.
RECAP_HOTEL_CODE_COLUMN = MLColumnNaming.recap_column("0_page_de_connexion_id_code_h")
RECAP_HOTEL_NAME_ALIAS_COLUMNS = frozenset(
    {MLColumnNaming.recap_column("0_page_de_connexion_id_nom_de_l_hotel")}
)
RECAP_IDENTITY_COLUMNS = frozenset(
    {RECAP_HOTEL_CODE_COLUMN, *RECAP_HOTEL_NAME_ALIAS_COLUMNS}
)
RECAP_GEO_LATITUDE_COLUMN = MLColumnNaming.recap_column(
    "0_page_de_connexion_localisation_geo_latitude"
)
RECAP_GEO_LONGITUDE_COLUMN = MLColumnNaming.recap_column(
    "0_page_de_connexion_localisation_geo_longitude"
)
RECAP_GEO_ADDRESS_COLUMNS = (
    MLColumnNaming.recap_column(
        "0_page_de_connexion_localisation_geo_adresse_postale_1"
    ),
    MLColumnNaming.recap_column(
        "0_page_de_connexion_localisation_geo_adresse_postale_2"
    ),
)
RECAP_GEO_CITY_COLUMN = MLColumnNaming.recap_column(
    "0_page_de_connexion_localisation_geo_ville"
)
RECAP_GEO_COORDINATE_COLUMNS = frozenset(
    {RECAP_GEO_LATITUDE_COLUMN, RECAP_GEO_LONGITUDE_COLUMN}
)
RECAP_NON_FEATURE_COLUMNS = RECAP_IDENTITY_COLUMNS | RECAP_GEO_COORDINATE_COLUMNS


class RodRecapExtractor:
    """Lit ``RECAP DATA ROD`` et produit un DataFrame wide par ``hotel_id``."""

    HOTEL_COL_START = 11

    def __init__(
        self,
        recap_path: Path,
        identity_registry: HotelIdentityRegistry,
        output_path: Path | None = None,
    ) -> None:
        self.recap_path = Path(recap_path)
        self.identity_registry = identity_registry
        self.output_path = Path(output_path) if output_path else None

    def _resolve_hotel_id(self, recap_column: str) -> str | None:
        if recap_column in RECAP_COLUMN_TO_HOTEL_ID:
            return RECAP_COLUMN_TO_HOTEL_ID[recap_column]
        resolved = self.identity_registry.resolve("rod", recap_column)
        if resolved:
            return resolved
        resolved = self.identity_registry.resolve("any", recap_column)
        return resolved

    @staticmethod
    def _slug(text: str) -> str:
        text = MLColumnNaming.fold_accents(str(text)).strip().lower()
        text = text.replace("%", "pct")
        text = re.sub(r"[^a-z0-9]+", "_", text)
        return re.sub(r"_+", "_", text).strip("_")[:80]

    @staticmethod
    def _normalize_raw(value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            v = value.strip()
            if v.upper() in {"OUI", "YES", "X"}:
                return 1
            if v.upper() in {"NON", "NO"}:
                return 0
            if v.upper() in MISSING_TOKENS or v == "-":
                return None
            if v.endswith("%"):
                try:
                    return float(v.replace("%", "").replace(",", ".")) / 100.0
                except ValueError:
                    return None
            if re.fullmatch(r"-?\d+,\d+", v):
                try:
                    return float(v.replace(",", "."))
                except ValueError:
                    return None
        if isinstance(value, (int, float)):
            return float(value)
        return value

    def extract_long(self) -> pd.DataFrame:
        if not self.recap_path.exists():
            raise FileNotFoundError(self.recap_path)

        wb = openpyxl.load_workbook(self.recap_path, data_only=True, read_only=True)
        ws = wb["RECAP DATA ROD"]
        hotel_columns: list[tuple[int, str, str | None]] = []
        for col in range(self.HOTEL_COL_START, ws.max_column + 1):
            header = ws.cell(3, col).value
            if not header:
                continue
            hotel_id = self._resolve_hotel_id(str(header).strip())
            hotel_columns.append((col, str(header).strip(), hotel_id))

        rows: list[dict[str, Any]] = []
        current_etape = ""
        current_sous = ""
        for row_idx in range(4, ws.max_row + 1):
            etape = ws.cell(row_idx, 2).value
            sous = ws.cell(row_idx, 3).value
            data_label = ws.cell(row_idx, 4).value
            if etape:
                current_etape = str(etape).strip()
            if sous:
                current_sous = str(sous).strip()
            if not data_label:
                continue

            field_key = self._slug(f"{current_etape}_{current_sous}_{data_label}")
            field_type_hint = self._infer_type_hint(str(data_label), row_idx)

            for col, recap_name, hotel_id in hotel_columns:
                if not hotel_id:
                    continue
                raw = self._normalize_raw(ws.cell(row_idx, col).value)
                rows.append(
                    {
                        "hotel_id": hotel_id,
                        "recap_column": recap_name,
                        "row": row_idx,
                        "etape": current_etape,
                        "sous_etape": current_sous,
                        "data_label": str(data_label).strip(),
                        "field_key": field_key,
                        "field_type_hint": field_type_hint,
                        "raw_value": raw,
                    }
                )
        wb.close()
        return pd.DataFrame(rows)

    @staticmethod
    def _infer_type_hint(label: str, row: int) -> str:
        label_l = label.lower()
        if "%" in label or "pct" in label_l:
            return "numeric_pct"
        if any(
            kw in label_l
            for kw in (
                "bar",
                "restaurant",
                "spa",
                "piscine",
                "minibar",
                "room-service",
                "corner",
                "réfrigér",
                "micro-ondes",
                "fontaine",
                "machine",
                "bouilloire",
                "vitrine",
                "top 1",
                "réception",
                "distributeur",
                "frigo",
                "caisse",
                "armoire",
            )
        ):
            return "boolean"
        if any(
            kw in label_l
            for kw in (
                "nb.",
                "nombre",
                "# ",
                "to ",
                "taux",
                "mètres",
                "longitude",
                "latitude",
                "chambres",
                "guests",
                "panier",
                "annee",
            )
        ):
            return "numeric"
        return "categorical"

    def extract_wide(self) -> pd.DataFrame:
        """Pivot long → wide avec préfixe ``d_recap_*`` + métadonnées de champ."""
        long_df = self.extract_long()
        if long_df.empty:
            return pd.DataFrame(columns=["hotel_code"])

        # field_key -> première ligne Excel ; suffixe _rN seulement si le libellé
        # réapparaît sur une autre ligne (pas entre hôtels sur la même ligne).
        seen: dict[str, int] = {}
        col_names: list[str] = []
        for _, row in long_df.iterrows():
            field_key = row["field_key"]
            excel_row = row["row"]
            base = f"recap_{field_key}"
            first_row = seen.get(field_key)
            if first_row is None:
                seen[field_key] = excel_row
            elif first_row != excel_row:
                field_key = f"{field_key}_r{excel_row}"
            col_names.append(MLColumnNaming.recap_column(field_key))

        long_df = long_df.copy()
        long_df["feature_column"] = col_names

        wide = (
            long_df.pivot_table(
                index="hotel_id",
                columns="feature_column",
                values="raw_value",
                aggfunc="first",
            )
            .reset_index()
        )
        wide.columns.name = None
        if RECAP_HOTEL_CODE_COLUMN in wide.columns:
            wide["code_h"] = wide[RECAP_HOTEL_CODE_COLUMN].astype(str).str.strip()
            wide.loc[wide["code_h"].isin(("", "nan", "None")), "code_h"] = None
        drop_cols = [c for c in RECAP_IDENTITY_COLUMNS if c in wide.columns]
        if drop_cols:
            wide = wide.drop(columns=drop_cols)

        schema = (
            long_df[
                ["feature_column", "field_key", "data_label", "field_type_hint", "row"]
            ]
            .drop_duplicates(subset=["feature_column"])
            .rename(columns={"feature_column": "column"})
        )

        if self.output_path:
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            long_df.to_csv(self.output_path.with_suffix(".long.csv"), index=False)
            wide.to_csv(self.output_path.with_suffix(".wide.csv"), index=False)
            schema.to_json(
                self.output_path.with_suffix(".schema.json"),
                orient="records",
                force_ascii=False,
                indent=2,
            )
        return wide