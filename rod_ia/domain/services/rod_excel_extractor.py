"""Extraction des constantes ROD depuis les classeurs Excel officiels."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl


class RodExcelExtractor:
    """Lit SIMULATEUR * + coûts et produit ``rod_reference.json`` traçable."""

    CONCEPTS = ("SIMPLY", "LIBERTY", "CONNECTED")

    # Lignes coûts SIMULATEUR * (label_col B, qty C, capex E, monthly H)
    TECHNO_ROWS = (
        (147, "scanner", "Scanner"),
        (148, "vitrine", "Vitrine"),
        (149, "licence", "Licence"),
        (150, "frais_os", "Frais OS"),
    )
    ANNEXES_ROWS = (
        (155, "elec_scanner", "Elec. scanner"),
        (156, "elec_vitrine", "Elec. vitrine"),
        (157, "staff", "Staff"),
    )

    def __init__(self, simulator_xlsx: Path, output_path: Path) -> None:
        self.simulator_xlsx = Path(simulator_xlsx)
        self.output_path = Path(output_path)

    def extract(self) -> dict[str, Any]:
        if not self.simulator_xlsx.exists():
            raise FileNotFoundError(self.simulator_xlsx)

        wb = openpyxl.load_workbook(self.simulator_xlsx, data_only=True, read_only=True)
        concepts: dict[str, Any] = {}
        for concept in self.CONCEPTS:
            sheet = f"SIMULATEUR {concept}"
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            pivot_m_lin = self._num(ws["F9"].value)
            concepts[concept] = {
                "pivot_nb_chambres": self._num(ws["C9"].value),
                "pivot_guests_per_chambre": self._num(ws["C10"].value),
                "pivot_to": self._num(ws["C11"].value),
                "pivot_m_lin": pivot_m_lin,
                "mix_fb": self._num(ws["I9"].value),
                "mix_nf": self._num(ws["I10"].value),
                "margin_fb_pct": self._num(ws["J9"].value),
                "margin_nf_pct": self._num(ws["J10"].value),
                "base_monthly_sales": self._num(ws["C19"].value),
                "base_monthly_ca_fb": self._num(ws["E120"].value, ws["E34"].value),
                "base_monthly_ca_nf": self._num(ws["E121"].value, ws["E35"].value),
                "monthly_cost_total": self._num(ws["H168"].value),
                "marge_nette_mensuelle_pilote": self._num(ws["E176"].value),
                "cost_lines": self._extract_cost_lines(ws, pivot_m_lin, concept),
            }
            fb = concepts[concept]["base_monthly_ca_fb"]
            nf = concepts[concept]["base_monthly_ca_nf"]
            concepts[concept]["base_monthly_ca"] = fb + nf
            concepts[concept]["techno_monthly"] = sum(
                line["monthly_unit"] * line["qty_default"]
                for line in concepts[concept]["cost_lines"]["techno"]
            )
            concepts[concept]["annexes_monthly"] = sum(
                line["monthly_unit"] * line["qty_default"]
                for line in concepts[concept]["cost_lines"]["annexes"]
            )
            ag = concepts[concept]["cost_lines"]["agencement"]
            concepts[concept]["agencement_per_m"] = ag["capex_per_m"]
            concepts[concept]["amort_months"] = ag["amort_months"]

        impact_to = self._extract_impact_to(wb)
        wb.close()

        payload = {
            "_source": str(self.simulator_xlsx.name),
            "_note": "Extrait automatiquement — cellules SIMULATEUR * (mois moyen pilote)",
            "impact_to": impact_to,
            "concepts": concepts,
        }
        self._merge_cost_defaults(payload)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.output_path.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return payload

    def _extract_cost_lines(
        self, ws: Any, pivot_m_lin: float, concept: str
    ) -> dict[str, Any]:
        techno: list[dict] = []
        for row, line_id, label in self._techno_rows_for(concept):
            raw_qty = ws.cell(row, 3).value
            qty = self._num(raw_qty, default=1.0 if raw_qty is None else 0.0)
            capex_unit = self._num(ws.cell(row, 5).value, default=0.0)
            monthly_unit = self._num(ws.cell(row, 8).value, default=0.0)
            techno.append(
                {
                    "id": line_id,
                    "label": str(ws.cell(row, 2).value or label),
                    "group": "techno",
                    "qty_default": qty,
                    "capex_unit": capex_unit,
                    "monthly_unit": monthly_unit,
                    "amort_months": 60.0,
                }
            )

        annexes: list[dict] = []
        for row, line_id, label in self.ANNEXES_ROWS:
            raw_qty = ws.cell(row, 3).value
            qty = self._num(raw_qty, default=1.0 if raw_qty is None else 0.0)
            capex_unit = self._num(ws.cell(row, 5).value, default=0.0)
            monthly_unit = self._num(ws.cell(row, 8).value, default=0.0)
            annexes.append(
                {
                    "id": line_id,
                    "label": str(ws.cell(row, 2).value or label),
                    "group": "annexes",
                    "qty_default": qty if qty else 1.0,
                    "capex_unit": capex_unit,
                    "monthly_unit": monthly_unit,
                    "amort_months": 60.0,
                }
            )

        # Agencement : cellule E166 / C166 (capex total) et H166 (mensuel)
        agencement_capex_total = self._num(ws["E166"].value, default=0.0)
        agencement_monthly = self._num(ws["H166"].value, default=0.0)
        ag_qty = self._num(ws["C166"].value, pivot_m_lin, default=pivot_m_lin or 1.0)
        capex_per_m = agencement_capex_total / ag_qty if ag_qty else 0.0
        amort = 84.0
        if agencement_monthly > 0 and agencement_capex_total > 0:
            amort = agencement_capex_total / agencement_monthly

        return {
            "techno": techno,
            "annexes": annexes,
            "agencement": {
                "label": str(ws["B166"].value or "Agencement"),
                "capex_per_m": capex_per_m,
                "amort_months": amort,
                "qty_default": ag_qty,
            },
        }

    @staticmethod
    def _techno_rows_for(concept: str) -> tuple[tuple[int, str, str], ...]:
        if concept == "LIBERTY":
            return (
                (147, "caisse", "Caisse"),
                (148, "vitrine", "Vitrine"),
                (149, "licence", "Licence"),
                (150, "frais_os", "Frais OS"),
            )
        if concept == "CONNECTED":
            return (
                (147, "frigo_froid", "Frigo froid"),
                (148, "frigo_ambiant", "Frigo amb."),
                (149, "licence", "Licence"),
                (150, "frais_os", "Frais OS"),
            )
        return RodExcelExtractor.TECHNO_ROWS

    @staticmethod
    def _extract_impact_to(wb: Any) -> dict[str, float]:
        if "REVENUS - IMPACT TO" not in wb.sheetnames:
            return {"ht_per_0_01_to": 9.233974, "ttc_per_0_01_to": 10.403846}
        ws = wb["REVENUS - IMPACT TO"]
        return {
            "ht_per_0_01_to": RodExcelExtractor._num(ws["F12"].value, default=9.233974),
            "ttc_per_0_01_to": RodExcelExtractor._num(ws["I12"].value, default=10.403846),
        }

    def _merge_cost_defaults(self, payload: dict[str, Any]) -> None:
        demo_path = self.output_path.parent / "rod_reference_demo.json"
        if not demo_path.exists():
            return
        demo = json.loads(demo_path.read_text(encoding="utf-8"))
        for concept, data in payload.get("concepts", {}).items():
            ref = (demo.get("concepts") or {}).get(concept, {})
            for key in ("fixed_capex",):
                if key not in data and key in ref:
                    data[key] = ref[key]

    @staticmethod
    def _num(*values: Any, default: float = 0.0) -> float:
        for value in values:
            if value is None:
                continue
            if isinstance(value, str):
                token = value.strip().lower()
                if token in {"", "inclus", "n/a"}:
                    continue
            try:
                return float(value)
            except (TypeError, ValueError):
                continue
        return default


class BrandProjectionsExtractor:
    """Extrait statistiques marques / nb hôtels depuis l'Excel paramètres."""

    def __init__(self, parameters_xlsx: Path, output_path: Path) -> None:
        self.parameters_xlsx = Path(parameters_xlsx)
        self.output_path = Path(output_path)

    def extract(self) -> dict[str, Any]:
        wb = openpyxl.load_workbook(self.parameters_xlsx, data_only=True, read_only=True)
        brands: dict[str, Any] = {}
        if "NB CH 1" in wb.sheetnames:
            ws = wb["NB CH 1"]
            current_brand = None
            for row in ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=3, values_only=True):
                label, count, _ = (row + (None, None, None))[:3]
                if label and str(label).strip() in {
                    "IBIS BUDGET", "IBIS STYLES", "MERCURE", "NOVOTEL", "IBIS"
                }:
                    current_brand = str(label).strip()
                    brands[current_brand] = {"total_hotels": int(count or 0), "size_bands": {}}
                elif current_brand and label and count is not None:
                    try:
                        brands[current_brand]["size_bands"][str(label).strip()] = int(count)
                    except (TypeError, ValueError):
                        pass

        reco_rules: list[dict[str, Any]] = []
        if "REGLES POUR RECO DU CONCEPT" in wb.sheetnames:
            ws = wb["REGLES POUR RECO DU CONCEPT"]
            for row in ws.iter_rows(min_row=4, max_row=12, min_col=4, max_col=6, values_only=True):
                brand_code, nb_hotels, _ = (row + (None, None, None))[:3]
                if brand_code and nb_hotels is not None:
                    try:
                        reco_rules.append(
                            {"brand_code": str(brand_code), "nb_hotels": int(nb_hotels)}
                        )
                    except (TypeError, ValueError):
                        pass

        wb.close()
        payload = {
            "_source": str(self.parameters_xlsx.name),
            "brands": brands,
            "recommendation_rules_sample": reco_rules,
        }
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return payload